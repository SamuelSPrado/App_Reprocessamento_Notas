import json
from datetime import datetime

def salvar_registro_erro(record, file_name="registros_erro.txt"):
    with open(file_name, "a") as file:
        file.write(json.dumps(record, indent=4) + "\n")

def criar_excel_com_cabecalhos(excel_file_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registros"
    sheet.append([
        "executionDate", "ownerId", "invoiceOrderId",
        "endpointResponseMessage", "id", "Data_Recebido"
    ])
    workbook.save(excel_file_path)

def salvar_registro_sucesso(record, excel_file_path):
    from openpyxl import load_workbook
    from os.path import exists

    if not exists(excel_file_path):
        criar_excel_com_cabecalhos(excel_file_path)

    with load_workbook(excel_file_path) as workbook:
        sheet = workbook.active
        dados_sucesso = {
            "executionDate": record.get("executionDate"),
            "ownerId": record.get("ownerId"),
            "invoiceOrderId": record.get("invoiceOrderId"),
            "endpointResponseMessage": record.get("endpointResponseMessage"),
            "id": record.get("id"),
            "Data_Recebido": datetime.now().isoformat()
        }
        sheet.append([
            dados_sucesso["executionDate"],
            dados_sucesso["ownerId"],
            dados_sucesso["invoiceOrderId"],
            dados_sucesso["endpointResponseMessage"],
            dados_sucesso["id"],
            dados_sucesso["Data_Recebido"]
        ])
        workbook.save(excel_file_path)