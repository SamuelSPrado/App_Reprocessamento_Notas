import requests
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
import time

class RequestManager:
    def __init__(self, owner_id, log_function, excel_file_path=None):
        self.endpoint_url = "https://third-api.meep.cloud/api/protheus/order/send"
        self.owner_id = owner_id
        self.log_function = log_function
        self.excel_file_path = excel_file_path
        self.max_retries = 2  # Tentar pelo menos uma vez além da tentativa inicial
        self.retry_delay = 5  # em segundos

    def salvar_registro_erro(self, record: dict) -> None:
        """Salva registros de erro em um arquivo de texto."""
        with open("registros_erro.txt", "a") as file:
            file.write(json.dumps(record, indent=4) + "\n")

    def criar_excel_com_cabecalhos(self) -> None:
        """Cria um novo arquivo Excel com cabeçalhos se não existir."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registros"
        # Adicionar cabeçalhos
        sheet.append([
            "executionDate", "ownerId", "invoiceOrderId",
            "endpointResponseMessage", "id", "Data_Recebido"
        ])
        workbook.save(self.excel_file_path)

    def salvar_registro_sucesso(self, record: dict) -> None:
        """Salva registros de sucesso em um arquivo Excel."""
        if not os.path.exists(self.excel_file_path):
            self.criar_excel_com_cabecalhos()

        # Carregar o workbook existente
        workbook = load_workbook(self.excel_file_path)
        sheet = workbook.active

        # Preparar os dados de sucesso
        dados_sucesso = {
            "executionDate": record.get("executionDate"),
            "ownerId": record.get("ownerId"),
            "invoiceOrderId": record.get("invoiceOrderId"),
            "endpointResponseMessage": record.get("endpointResponseMessage"),
            "id": record.get("id"),
            "Data_Recebido": datetime.now().isoformat()
        }

        # Adicionar uma nova linha com os dados
        sheet.append([
            dados_sucesso["executionDate"],
            dados_sucesso["ownerId"],
            dados_sucesso["invoiceOrderId"],
            dados_sucesso["endpointResponseMessage"],
            dados_sucesso["id"],
            dados_sucesso["Data_Recebido"]
        ])

        # Salvar o workbook
        workbook.save(self.excel_file_path)

    def registrar_erro_geral(self, mensagem: str) -> None:
        """Registra erros gerais com horário."""
        erro = {
            "timestamp": datetime.now().isoformat(),
            "message": mensagem
        }
        with open("registros_erro.txt", "a") as file:
            file.write(json.dumps(erro, indent=4) + "\n")

    def processar_intervalo(self, start: datetime, end: datetime) -> None:
        """Processa um intervalo específico de tempo com mecanismo de retry para todos os erros."""
        data = {
            'ownerId': self.owner_id,
            'start': start.strftime("%Y-%m-%d %H:%M:%S"),
            'end': end.strftime("%Y-%m-%d %H:%M:%S")
        }

        # Contadores
        count_incluido_anteriomente = 0
        count_reprocessado_com_sucesso = 0

        for attempt in range(self.max_retries):
            try:
                response = requests.post(self.endpoint_url, json=data)

                # Verificando se a solicitação foi bem-sucedida
                if response.status_code == 200:
                    try:
                        # Analisando a resposta JSON
                        response_data = response.json()

                        # Verificação inicial do tipo de resposta
                        if isinstance(response_data, list):
                            self.log_function(f"Total de registros recebidos: {len(response_data)}")

                            # Iterando sobre os registros na lista
                            for record in response_data:
                                endpoint_response_message = record.get("endpointResponseMessage", "")

                                if endpoint_response_message == "Pedido incluído anteriormente! ":
                                    count_incluido_anteriomente += 1
                                    self.salvar_registro_sucesso(record)  # Salvar registro de sucesso
                                elif endpoint_response_message == "":
                                    count_reprocessado_com_sucesso += 1
                                    self.salvar_registro_sucesso(record)  # Salvar registro de sucesso
                                else:
                                    self.salvar_registro_erro(record)  # Salvar registro de erro

                            # Exibindo os resultados no console
                            self.log_function(f"{count_incluido_anteriomente} de pedido já incluído anteriormente!")
                            self.log_function(f"{count_reprocessado_com_sucesso} de pedido reprocessado com sucesso.")
                        else:
                            self.log_function("Resposta JSON não é uma lista conforme esperado.")
                            self.registrar_erro_geral(f"Resposta JSON não é uma lista conforme esperado: {response_data}")

                        # Se o processamento foi bem-sucedido, saia do loop de retry
                        break

                    except json.JSONDecodeError:
                        self.log_function("Erro ao decodificar a resposta JSON.")
                        self.registrar_erro_geral("Erro ao decodificar a resposta JSON.")
                        break  # Saia do loop de retry em caso de erro de decodificação
                else:
                    self.log_function(f"Erro na solicitação: {response.status_code}")
                    self.registrar_erro_geral(f"Erro na solicitação: {response.status_code} - {response.text}")

                    # Aguarde e tente novamente para qualquer status code de erro
                    self.log_function(f"Esperando {self.retry_delay} segundos antes de tentar novamente...")
                    time.sleep(self.retry_delay)

            except requests.exceptions.RequestException as e:
                self.log_function(f"Erro ao fazer a solicitação: {e}")
                self.registrar_erro_geral(f"Erro ao fazer a solicitação: {e}")

                # Aguarde e tente novamente em caso de erro de requisição
                self.log_function(f"Esperando {self.retry_delay} segundos antes de tentar novamente...")
                time.sleep(self.retry_delay)

        # Adiciona uma linha em branco após o processamento de cada intervalo
        self.log_function("")

    def processar_por_id(self, file_path):
        excel_file_path = f"Reprocessamento 2 {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        with open(file_path, "r") as file:
            ids = file.readlines()

        for id_venda in ids:
            id_venda = id_venda.strip()
            if not id_venda:
                continue

            url = f"{ENDPOINT_URL}/{id_venda}"
            try:
                response = requests.post(url, json={"ownerId": self.owner_id})
                response.raise_for_status()

                if response.status_code == 200:
                    record = response.json()
                    record["id"] = id_venda
                    salvar_registro_sucesso(record, excel_file_path)
                    self.log_callback(f"{id_venda} {datetime.now()} {response.status_code} {record.get('endpointResponseMessage', '')}")
                else:
                    erro = {
                        "id": id_venda,
                        "status_code": response.status_code,
                        "message": response.text
                    }
                    salvar_registro_erro(erro)
                    self.log_callback(f"{id_venda} {datetime.now()} {response.status_code} ERRO")
            except requests.exceptions.RequestException as e:
                erro = {
                    "id": id_venda,
                    "message": str(e)
                }
                salvar_registro_erro(erro)
                self.log_callback(f"{id_venda} {datetime.now()} ERRO: {e}")

            time.sleep(TIME_SLEEP_ID)